import os
import tempfile

from marker.providers.pdf import PdfProvider

css = '''
@page {
    size: A4 landscape;
    margin: 1.5cm;
}

table {
    width: 100%;
    border-collapse: collapse;
    break-inside: auto;
    font-size: 10pt;
}

tr {
    break-inside: avoid;
    page-break-inside: avoid;
}

td {
    border: 0.75pt solid #000;
    padding: 6pt;
}
'''


class SpreadSheetProvider(PdfProvider):
    def __init__(self, filepath: str, config=None):
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=f".pdf")
        self.temp_pdf_path = temp_pdf.name
        temp_pdf.close()

        # Convert XLSX to PDF
        try:
            self.convert_xlsx_to_pdf(filepath)
        except Exception as e:
            raise RuntimeError(f"Failed to convert {filepath} to PDF: {e}")

        # Initialize the PDF provider with the temp pdf path
        super().__init__(self.temp_pdf_path, config)

    def __del__(self):
        if os.path.exists(self.temp_pdf_path):
            os.remove(self.temp_pdf_path)

    def convert_xlsx_to_pdf(self, filepath: str):
        from weasyprint import CSS, HTML
        from openpyxl import load_workbook

        html = ""
        workbook = load_workbook(filepath)
        if workbook is not None:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                html += f'<div><h1>{sheet_name}</h1>' + self._excel_to_html_table(sheet) + '</div>'
        else:
            raise ValueError("Invalid XLSX file")

        # We convert the HTML into a PDF
        HTML(string=html).write_pdf(
            self.temp_pdf_path,
            stylesheets=[CSS(string=css), self.get_font_css()]
        )

    @staticmethod
    def _get_merged_cell_ranges(sheet):
        merged_info = {}
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_info[(min_row, min_col)] = {
                'rowspan': max_row - min_row + 1,
                'colspan': max_col - min_col + 1,
                'range': merged_range
            }
        return merged_info

    def _excel_to_html_table(self, sheet):
        merged_cells = self._get_merged_cell_ranges(sheet)

        html = f'<table>'

        # Track cells we should skip due to being part of a merge range
        skip_cells = set()

        for row_idx, row in enumerate(sheet.rows, 1):
            html += '<tr>'
            for col_idx, cell in enumerate(row, 1):
                if (row_idx, col_idx) in skip_cells:
                    continue

                # Check if this cell is the start of a merged range
                merge_info = merged_cells.get((row_idx, col_idx))
                if merge_info:
                    # Add cells to skip
                    for r in range(row_idx, row_idx + merge_info['rowspan']):
                        for c in range(col_idx, col_idx + merge_info['colspan']):
                            if (r, c) != (row_idx, col_idx):
                                skip_cells.add((r, c))

                    # Add merged cell with rowspan/colspan
                    value = cell.value if cell.value is not None else ''
                    html += f'<td rowspan="{merge_info["rowspan"]}" colspan="{merge_info["colspan"]}">{value}'
                else:
                    # Regular cell
                    value = cell.value if cell.value is not None else ''
                    html += f'<td>{value}'

                html += '</td>'
            html += '</tr>'
        html += '</table>'
        return html
